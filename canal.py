#coding:utf-8
import sys
import weakref
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
import collections
import os
import re
import enum
import argparse
import math

def parse_args(arguments):
	p = argparse.ArgumentParser('This is script to read an collect data from cycle vtu calculation')
	pg = p.add_mutually_exclusive_group()
	pg.add_argument('-n', type=int, default=10, help='number of extracted records. default is 10')
	pg.add_argument('-l', type=int, nargs='+', help='specified list of node numbers')
	pf = p.add_mutually_exclusive_group()
	pf.add_argument('-c', action='store_false', help='collapse history history and insert new type cycles string in cell')
	pf.add_argument('-a', action='store_true', help='add additional infromation into type cycles column')
	p.add_argument('--limit', type=float, default=1E-8, help='set the limit of extracted types of cycle. default is 1E-8')
	p.add_argument('--outfile', type=str, default='table.xlsx', help='name of output file, default = table.xlsx')
	r = p.parse_args(arguments)
	return r


class ChildMixin():
	def __init__(self, parent, parent_type, key=None):
		if parent is None:
			self._parent = None
		elif isinstance(parent, parent_type):
			self._parent=weakref.ref(parent)
			if issubclass(parent_type, collections.UserList):
				self._parent().append(self)
			elif issubclass(parent_type, collections.UserDict):
				self._parent()[key] = self
		else:
			raise ValueError
	
	@property
	def parent(self):
		if self._parent is None:
			return None
		else:
			return self._parent()

class CycleTypeRecord(ChildMixin):
	
	def __init__(self, parent, first_id, second_id, saf ,sfmax, sfmin, tmax, tmin, r, ndop, n, a):
		super().__init__(parent, CycleTypeTable)
		self._first_id = first_id
		self._second_id = second_id
		self._saf = saf
		self._sfmax = sfmax
		self._sfmin = sfmin
		self._tmax = tmax
		self._tmin = tmin
		self._r = r
		self._ndop = ndop
		self._n = n
		self._a = a
	
	@property
	def first_id(self):
		return self._first_id
	
	@first_id.setter
	def first_id(self, value):
		self._first_id = value
	
	@property
	def second_id(self):
		return self._second_id
		
	@second_id.setter
	def second_id(self, value):
		self._second_id = value
	
	@property
	def saf(self):
		return self._saf
	
	@property
	def sfmax(self):
		return self._sfmax
	
	@property
	def sfmin(self):
		return self._sfmin
		
	@property
	def tmax(self):
		return self._tmax
	
	@property
	def tmin(self):
		return self._tmin
		
	@property
	def r(self):
		return self._r
	
	@property
	def ndop(self):
		return self._ndop
	
	@property
	def n(self):
		return self._n
		
	@property
	def a(self):
		return self._a
	

class CycleTypeTable(collections.UserList, ChildMixin):
	OUTPUT_HEADER = "fid   sid  sfmax   sfmin   saf      tmin   tmax   r      ndop       n        a"
	OUTPUT_FORMAT = "{fid:>3} - {sid:<3} {sfmax:>8.2f}{sfmin:>8.2f}{saf:>8.2f}{tmin:>7.1f}{tmax:>7.1f}{r:>7.2f}{ndop:>10.0f}{n:>10.1f} {a:>11.4e}"
	def __init__(self, nodenum, parent=None):
		self._nodenum = nodenum
		super().__init__()
		super(collections.UserList, self).__init__(parent, CycleTypeManagerTable, nodenum)
	
	@property
	def nodenum(self):
		return self._nodenum
	
	def print_table(self, limit=1E-8):
		print(self.OUTPUT_HEADER)
		for i in self:
			if i.a >= limit:
				print(self.OUTPUT_FORMAT.format(fid=i.first_id, sid=i.second_id, saf=i.saf, n=i.n, a=i.a, sfmax=i.sfmax, sfmin=i.sfmin, tmin=i.tmin, tmax=i.tmax, r=i.r, ndop=i.ndop))

class CycleTypeManagerTable(collections.UserDict):
	class AccumulatedFatigueDamageFileLineContext(enum.Enum):
		FIND_NODE_NUM = 0
		FIND_NUM_COMPONENT_NUM = 1
		FIND_NODE_BM = 2
		READ_RECORD = 3
		
	def __init__(self, node_table, local_reduced_stress_manager_table, elastic_reduced_stress_manager_table):
		super().__init__()
		if node_table is None:
			self._node_table = None
		elif isinstance(node_table, NodeTable):
			self._node_table=weakref.ref(node_table)
		else:
			raise ValueError
	
		if local_reduced_stress_manager_table is None:
			self._local_reduced_stress_manager_table = None
		elif isinstance(local_reduced_stress_manager_table, LocalReducedStressManagerTable):
			self._local_reduced_stress_manager_table=weakref.ref(local_reduced_stress_manager_table)
		else:
			raise ValueError
		
		if elastic_reduced_stress_manager_table is None:
			self._elastic_reduced_stress_manager_table = None
		elif isinstance(elastic_reduced_stress_manager_table, ElasticReducedStressManagerTable):
			self._elastic_reduced_stress_manager_table=weakref.ref(elastic_reduced_stress_manager_table)
		else:
			raise ValueError
		
	@property
	def elastic_reduced_stress_manager_table(self):
		if self._elastic_reduced_stress_manager_table is None:
			return None
		else:
			return self._elastic_reduced_stress_manager_table()
			
	@property
	def node_table(self):
		if self._node_table is None:
			return None
		else:
			return self._node_table()
			
	@property
	def local_reduced_stress_manager_table(self):
		if self._local_reduced_stress_manager_table is None:
			return None
		else:
			return self._local_reduced_stress_manager_table()
			
	
	def parse_accumulated_fatigue_damage_file(self, file):
		current_context = self.AccumulatedFatigueDamageFileLineContext.FIND_NODE_NUM
		current_table = None
		current_nodenum = None
		nessesery_component = None
		necessery_base_moment = None
		find_node_num_pattern = re.compile('(?<=\>\sCalculation\snode:\s)\d+')
		find_node_component_pattern = re.compile('(?<=\>\sComponent\snumber:\s)\d+')
		find_node_basemoment_pattern = re.compile('(?<=\>\sBase\scalculated\smoment\sof\stime\s)\d+')
		with open(file, mode='r') as f:
			for line in f:
				if current_context == self.AccumulatedFatigueDamageFileLineContext.FIND_NODE_NUM:
					result = re.search(find_node_num_pattern, line)
					if result:
						current_nodenum = int(result.group(0))
						if self.node_table.get(current_nodenum):
							nessesery_component = self.node_table[current_nodenum].component
							necessery_base_moment = self.node_table[current_nodenum].base_moment
							current_table = CycleTypeTable(int(result.group(0)), self)
							current_context = self.AccumulatedFatigueDamageFileLineContext.FIND_NUM_COMPONENT_NUM
				elif current_context == self.AccumulatedFatigueDamageFileLineContext.FIND_NUM_COMPONENT_NUM:
					result = re.search(find_node_component_pattern, line)
					if result and nessesery_component == int(result.group(0)):
						current_context = self.AccumulatedFatigueDamageFileLineContext.FIND_NODE_BM
				elif current_context == self.AccumulatedFatigueDamageFileLineContext.FIND_NODE_BM:
					result = re.search(find_node_basemoment_pattern, line)
					if result and necessery_base_moment == int(result.group(0)):
						current_context = self.AccumulatedFatigueDamageFileLineContext.READ_RECORD
						header = 2
				elif current_context == self.AccumulatedFatigueDamageFileLineContext.READ_RECORD:
					if header > 0:
						header -= 1
					else:
						temp_list = line.replace(',','.').strip().split()
						try:
							CycleTypeRecord(current_table, int(temp_list[0]), int(temp_list[2]), float(temp_list[7]), float(temp_list[5]), float(temp_list[6]), float(temp_list[9]), float(temp_list[8]), float(temp_list[10]), float(temp_list[19]), float(temp_list[20]), float(temp_list[21]))
						except (ValueError, IndexError ) as vi:
							current_table.sort(key = lambda a: a.a, reverse=True)
							current_context = self.AccumulatedFatigueDamageFileLineContext.FIND_NODE_NUM
	
class LocalReducedStressRecord(ChildMixin):
	
	def __init__(self, num, parent= None, temp:float=20.0, si:float=0.0, sj:float=0.0, sk:float=0.0):
		self._num = num
		self._temp = temp
		self._list = [si, sj, sk, None, None, None]
		super().__init__(parent, LocalReducedStressTable, num)
	
	@property
	def num(self):
		return self._num
		
	@property
	def temp(self):
		return self._temp
		
	@property
	def si(self):
		return self._list[0]
	
	@property
	def sj(self):
		return self._list[1]
		
	@property
	def sk(self):
		return self._list[2]
		
	@property
	def sij(self):
		if self._list[3] is None:
			self._list[3] = self._list[0] - self._list[1]
		return self._list[3]
	
	@property
	def sjk(self):
		if self._list[4] is None:
			self._list[4] = self._list[1] - self._list[2]
		return self._list[4]
		
	@property
	def sik(self):
		if self._list[5] is None:
			self._list[5] = self._list[0] - self._list[2]
		return self._list[5]
	
	@property
	def vec(self):
		if self._list[3] is None:
			self._list[3] = self._list[0] - self._list[1]
		if self._list[4] is None:
			self._list[4] = self._list[1] - self._list[2]
		if self._list[5] is None:
			self._list[5] = self._list[0] - self._list[2]
		return self._list.copy()

class LocalReducedStressTable(collections.UserDict, ChildMixin):
	def __init__(self, nodenum, parent):
		self._nodenum = nodenum
		super().__init__()
		super(collections.UserDict, self).__init__(parent, LocalReducedStressManagerTable, nodenum)

	@property
	def nodenum(self):
		return self._nodenum
	
	def print_table(self):
		print('id        temp      sij       sjk       sik')
		for moment, m in self.items():
			print("{moment:<10}{temp:<10.1f}{sij:<10.2f}{sjk:<10.2f}{sik:<10.2f}".format(moment=moment, temp=m.temp, sij=m.sij, sjk=m.sjk, sik=m.sik))

			

class LocalReducedStressManagerTable(collections.UserDict):
	class LocalReducedStresFileLineContext(enum.Enum):
		FIND_NODE_NUM = 0
		FIND_NODE_BM = 1
		READ_RECORD = 2
		
	def __init__(self, node_table):
		super().__init__()
		if node_table is None:
			self._node_table = None
		elif isinstance(node_table, NodeTable):
			self._node_table=weakref.ref(node_table)
		else:
			raise ValueError
	
	@property
	def node_table(self):
		if self._node_table is None:
			return None
		else:
			return self._node_table()
	
	@property
	def length_of_tables(self):
		return len(max(self.values(), key=lambda a: len(a)))
	
	def parse_local_redused_stress_file(self, file, verbose=False):
		current_context = self.LocalReducedStresFileLineContext.FIND_NODE_NUM
		current_table = None
		current_nodenum = None
		necessery_base_moment = None
		find_node_num_pattern = re.compile('(?<=\>\sCalculation\snode\s)\d+')
		find_node_basemoment_pattern = re.compile('(?<=\>\>moment\s)\d+(?=\s-\>\scalculation\sresults\sTable)')
		start_header_passer = False
		with open(file, mode='r') as f:
			for line in f:
				if current_context == self.LocalReducedStresFileLineContext.FIND_NODE_BM:
					result = re.search(find_node_basemoment_pattern, line)
					if result and necessery_base_moment == int(result.group(0)):
						if verbose:
							print('Find necessery_base_moment = {}'.format(necessery_base_moment))
						current_context = self.LocalReducedStresFileLineContext.READ_RECORD
						start_header_passer = True
				elif current_context == self.LocalReducedStresFileLineContext.READ_RECORD:
					if not start_header_passer:
						try:
							temp_list = line.replace(',','.').strip().split()
							LocalReducedStressRecord(int(temp_list[0]), current_table, float(temp_list[1]), float(temp_list[5]), float(temp_list[6]), float(temp_list[7]))
						except (ValueError, IndexError):
							current_context = self.LocalReducedStresFileLineContext.FIND_NODE_NUM
					else:
						start_header_passer = False
				if current_context == self.LocalReducedStresFileLineContext.FIND_NODE_NUM:
					start_header_passer = False
					result = re.search(find_node_num_pattern, line)
					if result:
						current_nodenum = int(result.group(0))
						if self.node_table.get(current_nodenum):
							if verbose:
								print('Find node header = {}'.format(current_nodenum))
							necessery_base_moment = self.node_table[current_nodenum].base_moment
							current_table = LocalReducedStressTable(int(result.group(0)), self)
							current_context = self.LocalReducedStresFileLineContext.FIND_NODE_BM
						
class NodeRecord(ChildMixin):
	def __init__(self, num, parent=None, damage:float=0.0, base_moment:int=0, component:int=0):
		self._num = num
		self._damage = damage
		self._base_moment = base_moment
		self._component = component
		super().__init__(parent, NodeTable, num)

	@property
	def num(self):
		return self._num
	
	@property
	def damage(self):
		return self._damage
	
	@property
	def base_moment(self):
		return self._base_moment
	
	@property
	def component(self):	
		return self._component
	
	@damage.setter
	def damage(self, value:float):
		if self._damage == 0.0 and value >= 0.0:
			self._damage = value
		else:
			raise ValueError
		
	@base_moment.setter
	def base_moment(self, value:int):
		if self._base_moment == 0 and value >= 0:
			self._base_moment = value
		else:
			raise ValueError(value,self._base_moment)
			
	@component.setter
	def component(self, value:int):
		if self._component == 0 and value in (1, 2, 3):
			self._component = value
		else:
			raise ValueError
		
class NodeTable(collections.UserDict):
	OUTPUT_HEADER = 'nodenum   damage          bm    component'
	OUTPUT_FORMAT = "{node:<10}{damage:<10.5e}{bm:6}{c:>6}"
	def __init__(self):
		super().__init__()
		self._dindex = None
		
	def parse_base_moments(self, file):
		extract = lambda typ, line, sign: typ(line.split(sign)[1].strip())
		with open(file, mode='r') as f:
			for line in f: 
				if line.startswith('Calculation'):
					current_node = extract(int, line, ':')
					NodeRecord(current_node, self)
				elif line.startswith('a = '):
					self[current_node].damage = extract(float, line.replace(',','.').strip().strip('.'), '=')
				elif line.startswith('Base calculated moment of time'):
					self[current_node].base_moment = extract(int, line.replace(',',''), ':')
				elif line.startswith('reduced sterss component'):
					self[current_node].component = extract(int, line.replace('.',''), ':')
				else:
					pass
	
	def get_damage_index(self, num=None):
		if self._dindex is None:
			self._dindex = sorted(self.values(), key=lambda a: a.damage,reverse=True)
		if num is None:
			return self._dindex
		else:
			return self._dindex[:num]
	
	def print_table(self, limit=0, sort_by_damage=False):
		print(self.OUTPUT_HEADER)
		if sort_by_damage:
			ld = list(self.items())
			ld.sort(key=lambda a: a[1].damage,reverse=True)
		else:
			ld = self.items()
		if limit is None or limit <= 0:
			for nodenum, node in ld:
				print(self.OUTPUT_FORMAT.format(node=nodenum, damage=node.damage, bm=node.base_moment, c=node.component))
		else:
			for num, (nodenum, node) in enumerate(ld):
				if num >= limit:
					break
				print(self.OUTPUT_FORMAT.format(node=nodenum, damage=node.damage, bm=node.base_moment, c=node.component))
	
	def print_table_by_list(self, list_of_nodes, sort_by_damage=True):
		print(self.OUTPUT_HEADER)
		ld = []
		for item in list_of_nodes:
			ld.append(self[item])
		if sort_by_damage:
			ld.sort(key=lambda a: a[1].damage,reverse=True)
		for nodenum, node in ld:
			print(self.OUTPUT_FORMAT.format(node=nodenum, damage=node.damage, bm=node.base_moment, c=node.component)) 


class ElasticReducedStressRecord(ChildMixin):
	def __init__(self, parent, num, temp, rpe, nu, ksi=None, lb=None, lh=None, sll=0.0, sfl=0.0):
		super().__init__(parent, ElasticReducedStressTable, num)
		self._num = num
		self._temp = temp
		self._rpe = rpe
		self._nu = nu
		self._ksi = ksi
		self._lb = lb
		self._lh = lh
		self._sll = sll
		self._sfl = sfl
		self._rid = None

	@property
	def num(self):
		return self._num
	
	@property
	def temp(self):
		return self._temp
	
	@property
	def rpe(self):
		return self._rpe
	
	@property
	def nu(self):
		return self._nu
	
	@property
	def ksi(self):	
		return self._ksi
	
	@property
	def lb(self):
		return self._lb
	
	@property
	def lh(self):
		return self._lh
	
	@property
	def sll(self):
		return self._sll
	
	@property
	def sfl(self):
		return self._sfl
	
	@property
	def real_id(self):
		if self.parent:
			nn = self.parent.nodenum
			lt = self.parent.parent.local_reduced_stress_manager_table[nn]
			for key, values in lt.items():
				if math.isclose(self.temp, values.temp, abs_tol=1E-2) and math.isclose(self.sll,values.vec[2+self.parent.parent.node_table[nn].component],abs_tol=1E-4):
					self._rid = key
		return self._rid
					
					
	
class ElasticReducedStressTable(collections.UserDict, ChildMixin):
	def __init__(self, nodenum, parent):
		self._nodenum = nodenum
		super().__init__()
		super(collections.UserDict, self).__init__(parent, ElasticReducedStressManagerTable, nodenum)

	@property
	def nodenum(self):
		return self._nodenum
	
	
	def search_real_id(self, stress):
		for item in self.values():
			if math.isclose(item.sfl, stress,abs_tol=1E-4):
				return item.real_id
	
	def print_table(self):
		pass
		#print('id        temp      sij       sjk       sik')
		#for moment, m in self.items():
			#print("{moment:<10}{temp:<10.1f}{sij:<10.2f}{sjk:<10.2f}{sik:<10.2f}".format(moment=moment, temp=m.temp, sij=m.sij, sjk=m.sjk, sik=m.sik))
					
class ElasticReducedStressManagerTable(collections.UserDict):
	class ElasticReducedStressFileLineContext(enum.Enum):
		FIND_NODE_NUM = 0
		FIND_NUM_COMPONENT_NUM = 1
		FIND_NODE_BM = 2
		READ_RECORD = 3
		
	def __init__(self, node_table, local_reduced_stress_manager_table):
		super().__init__()
		if node_table is None:
			self._node_table = None
		elif isinstance(node_table, NodeTable):
			self._node_table=weakref.ref(node_table)
		else:
			raise ValueError
		
		if local_reduced_stress_manager_table is None:
			self._local_reduced_stress_manager_table = None
		elif isinstance(local_reduced_stress_manager_table,LocalReducedStressManagerTable):
			self._local_reduced_stress_manager_table = weakref.ref(local_reduced_stress_manager_table)
		else:
			raise ValueError
	
	@property
	def node_table(self):
		if self._node_table is None:
			return None
		else:
			return self._node_table()
	
	@property
	def local_reduced_stress_manager_table(self):
		if self._local_reduced_stress_manager_table is None:
			return None
		else:
			return self._local_reduced_stress_manager_table()
	
	def parse_elastic_reduced_stress_file(self, file):
		current_context = self.ElasticReducedStressFileLineContext.FIND_NODE_NUM
		current_table = None
		current_nodenum = None
		nessesery_component = None
		necessery_base_moment = None
		find_node_num_pattern = re.compile('(?<=\>\sCalculation\snode\s)\d+')
		find_node_component_pattern = re.compile('(?<=\>\sComponent\snumber\s)\d+')
		find_node_basemoment_pattern = re.compile('(?<=\>\sBase\scalculated\smoment\sof\stime\s)\d+')
		with open(file, mode='r') as f:
			for line in f:
				if current_context == self.ElasticReducedStressFileLineContext.FIND_NUM_COMPONENT_NUM:
					result = re.search(find_node_component_pattern, line)
					if result and nessesery_component == int(result.group(0)):
						current_context = self.ElasticReducedStressFileLineContext.FIND_NODE_BM
				elif current_context == self.ElasticReducedStressFileLineContext.FIND_NODE_BM:
					result = re.search(find_node_basemoment_pattern, line)
					if result and necessery_base_moment == int(result.group(0)):
						current_context = self.ElasticReducedStressFileLineContext.READ_RECORD
						header = 2
				elif current_context == self.ElasticReducedStressFileLineContext.READ_RECORD:
					if header > 0:
						header -= 1
					else:
						temp_list = line.replace(',', '.').strip().split()
						try:
							ElasticReducedStressRecord(current_table, num=int(temp_list[0]), temp=float(temp_list[1]), rpe=float(temp_list[2]), nu=float(temp_list[3]), sll=float(temp_list[-3]), sfl=float(temp_list[-1]))
						except (ValueError, IndexError ) as vi:
							current_context = self.ElasticReducedStressFileLineContext.FIND_NODE_NUM
				if current_context == self.ElasticReducedStressFileLineContext.FIND_NODE_NUM:
					result = re.search(find_node_num_pattern, line)
					if result:
						current_nodenum = int(result.group(0))
						if self.node_table.get(current_nodenum):
							nessesery_component = self.node_table[current_nodenum].component
							necessery_base_moment = self.node_table[current_nodenum].base_moment
							current_table = ElasticReducedStressTable(int(result.group(0)), self)
							current_context = self.ElasticReducedStressFileLineContext.FIND_NUM_COMPONENT_NUM

def save_in_workbook(manager_table, necessery_nodes=None, worksheet_name='ma', limit=1E-8, is_expanded=True, additional =False):
	mb = openpyxl.Workbook()
	chwidth = (("Тип цикла", 12),
	          ("σFmax",     8.25),
	          ("σFmin",     8.25),
	          ("σaF",       8.25),
	          ("Tmin",      8.25),
	          ("Tmax",      8.25),
	          ("r",         8),
	          ("[N]",       11),
	          ("N",         9),
	          ("a",         10))
	font = Font(name='Times New Roman', size=12)
	thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
	cent_alignment = Alignment(horizontal="center",
                           vertical="center",
                           wrap_text=True)
	if necessery_nodes is None:
		necessery_nodes = manager_table.keys()
	mlen = manager_table.local_reduced_stress_manager_table.length_of_tables
	for node in necessery_nodes:
		is_empty = True
		sheet_name = '{}n'.format(node)
		ms = mb.create_sheet(sheet_name)
		ct = manager_table[node]
		for rownum, item in enumerate(filter(lambda a: a.a > limit, ct), 2):
			if is_expanded and (item.first_id > mlen or item.second_id > mlen):
				new_first_id = manager_table.elastic_reduced_stress_manager_table[node].search_real_id(item.sfmax)
				new_second_id = manager_table.elastic_reduced_stress_manager_table[node].search_real_id(item.sfmin)
				if additional and new_first_id and new_second_id:	
					ms.cell(row=rownum, column=1).value = "{}-{} ({}-{})".format(item.first_id, item.second_id, new_first_id, new_second_id)
				else:
					ms.cell(row=rownum, column=1).value = "{}-{}".format(new_first_id, new_second_id)
				new_first_id = None
				new_second_id = None
			else:
				ms.cell(row=rownum, column=1).value = "{}-{}".format(item.first_id, item.second_id)
			ms.cell(row=rownum, column=2).value = item.sfmax   
			ms.cell(row=rownum, column=3).value = item.sfmin
			ms.cell(row=rownum, column=4).value = item.saf
			ms.cell(row=rownum, column=5).value = item.tmin
			ms.cell(row=rownum, column=6).value = item.tmax
			ms.cell(row=rownum, column=7).value = item.r
			ms.cell(row=rownum, column=8).value = item.ndop
			ms.cell(row=rownum, column=9).value = item.n
			ms.cell(row=rownum, column=10).value = item.a
			is_empty = False
		if manager_table.node_table[node].damage >= 1.0:
			ms.sheet_properties.tabColor = openpyxl.styles.colors.Color('FF0000')
		if not is_empty:
			if openpyxl.__version__[0] == '3':
				for cnum, (chvalue, cwidth) in enumerate(chwidth, 1):
					hcell = ms.cell(row=1, column=cnum)
					hcell.value = chvalue
					ms.column_dimensions[hcell.column_letter].width = cwidth
				ms.cell(row=rownum+1, column=10).value = "=SUM(J2:J{})".format(rownum)
				mc = ms.merge_cells(start_row=rownum+1, start_column=1, end_row=rownum+1, end_column=9)
				ms.cell(row=rownum+1, column=1).value = "Итоговая накопленная усталостная поврежденность"
				ms.cell(row=rownum+1, column=1).border = thin_border
				ms.cell(row=rownum+1, column=1).alignment = cent_alignment
				ms.cell(row=rownum+1, column=1).font = font
				ms.cell(row=rownum+1, column=10).border = thin_border
				ms.cell(row=rownum+1, column=10).alignment = cent_alignment
				ms.cell(row=rownum+1, column=10).font = font
				for row in ms['A1:J{}'.format(rownum)]:
					for cell in row:
						cell.border = thin_border
						cell.alignment = cent_alignment
						cell.font = font
						if cell.column_letter in 'BCDEFH':
							cell.number_format = '0'
						elif cell.column_letter == 'G':
							cell.number_format = '0.00'
						elif cell.column_letter == 'I':
							cell.number_format = '0.0'
						elif cell.column_letter == 'J':
							cell.number_format = '0.0E+0'
			elif openpyxl.__version__[0] == '2':
				for cnum, (chvalue, cwidth) in enumerate(chwidth, 1):
					hcell = ms.cell(row=1, column=cnum)
					hcell.value = chvalue
					ms.column_dimensions[hcell.column].width = cwidth
				ms.cell(row=rownum+1, column=10).value = "=SUM(J2:J{})".format(rownum)
				ms.merge_cells(start_row=rownum+1, start_column=1, end_row=rownum+1, end_column=9)
				ms.cell(row=rownum+1, column=1).value = "Итоговая накопленная усталостная поврежденность"
				for row in ms['A1:J{}'.format(rownum+1)]:
					for cell in row:
						cell.border = thin_border
						cell.alignment = cent_alignment
						cell.font = font
						if cell.column in 'BCDEFH':
							cell.number_format = '0'
						elif cell.column == 'G':
							cell.number_format = '0.00'
						elif cell.column == 'I':
							cell.number_format = '0.0'
						elif cell.column == 'J':
							cell.number_format = '0.0E+0'
	try:
		mb.save("{}".format(worksheet_name))
	except PermissionError:
		print('ERROR: Please close the excel file')
	
		
def main():
	args = parse_args(sys.argv[1:])
	if not args.l:
		nnodes = args.n
	else:
		nnodes = None
	nt = NodeTable()
	nt.parse_base_moments(max(list(filter(lambda a: a.startswith('BaseMoments'), os.listdir())), key=os.path.getctime))
	if nnodes:
		nt.print_table(sort_by_damage=True, limit=nnodes)
	else:
		nt.print_table_by_list(args.l)
	lmt = LocalReducedStressManagerTable(nt)
	lmt.parse_local_redused_stress_file(max(list(filter(lambda a: a.startswith('Report (Local Reduced Stress)'), os.listdir())), key=os.path.getctime))
	
	emt = ElasticReducedStressManagerTable(nt, lmt)
	emt.parse_elastic_reduced_stress_file(max(list(filter(lambda a: a.startswith('Report (Elastic Reduced Stress)'), os.listdir())), key=os.path.getctime))
	
	ctt = CycleTypeManagerTable(nt, lmt, emt)
	ctt.parse_accumulated_fatigue_damage_file(max(list(filter(lambda a: a.startswith('Report (Accumulated Fatigue Damage)'), os.listdir())), key=os.path.getctime))
	if nnodes:
		nn = list(map(lambda a: a.num, nt.get_damage_index(nnodes)))
	else:
		nn = args.l
	save_in_workbook(ctt, nn, args.outfile, args.limit, args.c or args.a, args.a)
	
	
if __name__ == "__main__":
	main()